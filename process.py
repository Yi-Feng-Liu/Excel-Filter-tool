import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, numbers
import copy
import time


class Judge_Metabolic_Syndrome:
    def __init__(self, io, select_years, save_sheet_name, save_file_path, from_summary=False):
        self.io = io
        self.dst_worksheet = save_sheet_name
        self.select_years = select_years
        self.save_file_path = save_file_path
        self.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.font = Font(name='Calibri', color='FF0000')
        self.font_type = Font(name='Calibri')
        self.from_summary = from_summary
        self.gender_dict = {'gender': 5}
        self.column_dict={
            'waistline':9,
            'systolic':10,
            'diastolic':11,
            'glucose':12,
            'triglycerides':14,
            'hdlc':15
        }
        self.standard_dict={
            'waistline': 90,
            'systolic': 130,
            'diastolic': 85,
            'glucose':100,
            'triglycerides':150,
            'hdlc': 40
        }
        # self.main_procesdure()
        

    def change_date_time(self, worksheet, number_of_column):
        """Remove hours:minute:second format of the datetime 

        Args:
            worksheet : excel worksheet
            number_of_column : If column name is G, the number of column is 7, etc.

        Returns:
            _worksheet: 
        """
        # 變更時間格式
        for row in worksheet.iter_rows(min_row=2, min_col=number_of_column, max_col=number_of_column):
            for cell in row:
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
        return worksheet


    def place_center(self, worksheet):
        align = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = align
        return worksheet


    def set_specific_column_format(self, worksheet:str, eng_column:str, width=20):
        """set the specific column format.

        Args:
            worksheet (str): Excel worksheet
            eng_column (str): like 'A' or 'G' column
        """
        worksheet.column_dimensions[eng_column].width = width
        worksheet[f'{eng_column}1'].fill = self.fill
        worksheet[f'{eng_column}1'].font = self.font
        return worksheet


    def change_font_color_format(self, cell):
        """Change font color 

        Args:
            cell : the cell coordinate
        """
        cell.font = self.font
    
    def change_font_type_format(self, worksheet):
        """Change font type 

        Args:
            cell : the cell coordinate
        """
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                # 設置儲存格的字體
                cell.font = self.font_type
        return worksheet

    def label_over_standard(self, worksheet):
        """Use to label the cell, if cell's value exceed the standard

        Args:
            worksheet: the excel work sheet 

        Returns:
            worksheet
        """
        worksheet = self.change_font_type_format(worksheet)

        for row in worksheet.iter_rows(min_row=2):
            people_name = row[1]
            gender = row[self.gender_dict['gender']] 
            over_standard_cnt = 0
            if gender.value == '男':
                for key, value in self.column_dict.items():
                    if row[value].value is None:
                        continue
                    elif isinstance(row[value].value, str):
                        row[value].value = float(row[value].value)
                        if key == 'hdlc' and row[value].value < self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if key != 'hdlc' and row[value].value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)

            elif gender.value == '女':
                for key, value in self.column_dict.items():
                    if row[value].value is None:
                        continue
                    elif isinstance(row[value].value, str):
                        row[value].value = float(row[value].value.split('(')[0])
                        if key == 'waistline' and row[value].value >= self.standard_dict[key]-10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if key =='hdlc' and row[value].value < self.standard_dict[key]+10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if key != 'hdlc' and row[value].value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)
        return worksheet
    

    def copy_title_format(self, ws1, ws2):
        for row in ws1.iter_rows(min_row=1, max_row=1):
            for cell in row:
                # copy cell
                ws2[cell.coordinate].font = copy.copy(cell.font)
                ws2[cell.coordinate].fill = copy.copy(cell.fill)
                ws2[cell.coordinate].border = copy.copy(cell.border)
                ws2[cell.coordinate].alignment = copy.copy(cell.alignment)
                ws2[cell.coordinate].number_format = copy.copy(cell.number_format)
                ws2.row_dimensions[cell.row].height = copy.copy(ws1.row_dimensions[cell.row].height)
                ws2.column_dimensions[cell.column_letter].width = copy.copy(ws1.column_dimensions[cell.column_letter].width)
        return ws2
    
    def copy_format_from_sheet1(self):
        """Copy the original sheet header format to specific sheet

        Including cell's fill, font, color, alignment, dimensions.
        """
        if self.from_summary==False:
            workbook = openpyxl.load_workbook(self.io)
        else:
            workbook = openpyxl.load_workbook(self.save_file_path)

        ws1 = workbook['健檢資料']
        ws2 = workbook[self.dst_worksheet]
        # copy format sheet1 header to sheet2 header
        ws2 = self.copy_title_format(ws1=ws1, ws2=ws2)
        ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='U')
        if self.from_summary==True:
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='V')
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='W')
        ws2 = self.change_date_time(worksheet=ws2, number_of_column=7)
        ws2 = self.place_center(worksheet=ws2)

        # label_over_standard worksheet
        # ws1 = self.label_over_standard(worksheet=ws1)
        # only process new sheet
        ws2 = self.label_over_standard(worksheet=ws2)

        if self.from_summary==False:
            workbook.save(self.io)
        else:
            workbook.save(self.save_file_path)
        print("saved")


    def read_file(self):
        if self.from_summary==False:
            df = pd.read_excel(self.io, sheet_name='健檢資料', engine='openpyxl')
        else:
            df = pd.read_excel(self.save_file_path, sheet_name='健檢資料', engine='openpyxl')
        return df
    
    
    def process_Metabolic_Syndrome(self, df:pd.DataFrame):
        """Filter file and add new column named 超過標準數,

        Args:
            df (pd.DataFrame): Original dataframe
        Returns:
            dataframe
        """
        df['超過標準數'] = 0
        df = df[df['年度代碼'].str.startswith(self.select_years)]
             
        for i in range(len(df.index)):
            gender = df.iloc[i, self.gender_dict['gender']] 
            over_standard_cnt = 0
            if gender == '男':
                for key, value in self.column_dict.items():
                    df_value = df.iloc[i, value]
                    if pd.isna(df_value) or len(df_value)==0:
                        continue
                    elif isinstance(df_value, str):
                        df_value = float(df_value)
                        if key == 'hdlc' and df_value < self.standard_dict[key]:
                            over_standard_cnt += 1
                        if key != 'hdlc' and df_value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                    df.iloc[i, len(df.columns)-1] = over_standard_cnt
                
            elif gender == '女':
                for key, value in self.column_dict.items():
                    df_value = df.iloc[i, value]
                    if pd.isna(df_value) or len(df_value)==0:
                        continue
                    elif isinstance(df_value, str):
                        df_value = float(df_value.split('(')[0])
                        if key == 'waistline' and df_value >= self.standard_dict[key]-10:
                            over_standard_cnt += 1
                        if key =='hdlc' and df_value < self.standard_dict[key]+10:
                            over_standard_cnt += 1
                        if key != 'hdlc' and df_value >= self.standard_dict[key]:
                            over_standard_cnt += 1   
                df.iloc[i, len(df.columns)-1] = over_standard_cnt
        df = df.sort_values(by=['超過標準數'], ascending=False)
        
        return df
        
    
    def save_file_and_copy_title(self, df):
        # 建立一個新的 ExcelWriter 物件
        if self.from_summary == False:
            writer = pd.ExcelWriter(self.io, mode='a', engine='openpyxl', if_sheet_exists='replace')
        else:
            writer = pd.ExcelWriter(self.save_file_path, mode='a', engine='openpyxl', if_sheet_exists='replace')
        df.to_excel(writer, sheet_name=self.dst_worksheet, index=False)
        writer.close() 


    def main_procesdure(self):
        df = self.read_file()
        df = self.process_Metabolic_Syndrome(df)
        self.save_file_and_copy_title(df)
        self.copy_format_from_sheet1()


class Metabolic_Syndrome_From_Summary(Judge_Metabolic_Syndrome):
    def __init__(self, io, select_years, save_sheet_name, save_file_path, from_summary=True):
        super().__init__(io, select_years, save_sheet_name, save_file_path)
        self.io = io
        self.select_years = select_years
        self.save_sheet_name = save_sheet_name
        self.save_file_path = save_file_path
        self.from_summary = from_summary
        self.main_procesdure()
        

    def append_column(self):
        self.df2 = pd.read_excel(self.save_file_path, sheet_name='健檢資料', engine='openpyxl')
        self.goal_column_name = self.df2.columns.tolist()
        if 'SGPT' not in self.df2.columns:
            self.df2['SGPT'] = 0
        if 'SGOT' not in self.df2.columns:
            self.df2['SGOT'] = 0    
        return self.df2.columns.tolist()
    
    def change_column_name(self):
        self.df = pd.read_excel(self.io, engine='openpyxl')
        self.goal_column_name = self.append_column()
        self.df['年度代碼'] = '111年度體檢'
        self.df['部門代號'] = 'X001'
        self.df['健檢過程備註說明'] = ''
        self.df['出生年月日'] = '2023/1/1'
        self.df = self.df.drop(labels=0, axis=0)
        
        specific_column = ['年度代碼', '姓名', '工/學號', '部門代號', '部門/科系', '性別', '出生年月日', '身高', '體重', '腰圍', '收縮壓', '舒張壓', 'AC飯前血糖', 'T-CHO總膽固醇', 'TG三酸甘油脂', 'HDL高密度脂蛋白', 'LDL低密度脂蛋白', '請問您過去一個月內是否有吸菸？', '既往病史', '健檢過程備註說明', 'SGOT血清麩酸草酸轉氨脢', 'SGPT血清麩酸丙銅轉氨脢']
        speific_df = self.df.copy()
        speific_df = speific_df[specific_column]
        speific_df_columns = speific_df.columns.to_list()

        for i in range(len(speific_df_columns)):
            speific_df.rename(columns={speific_df_columns[i]: self.goal_column_name[i]}, inplace=True)  

        # save the sheet 
        # self.process_Metabolic_Syndrome(io='test.xlsx', select_years=111, dst_worksheet='test1')
        # speific_df = speific_df.fillna(0)
        return speific_df
    
    def main_procesdure(self):
        speific_df = self.change_column_name()
        speific_df = self.process_Metabolic_Syndrome(speific_df)
        self.save_file_and_copy_title(speific_df)
        self.copy_format_from_sheet1()
        print('OK')
        
        
        
    
def main():
    # Judge_Metabolic_Syndrome('test.xlsx', '111', '工作表1', None)
    Metabolic_Syndrome_From_Summary('一般作業總表.xlsx', '111', 'test_data', save_file_path='test.xlsx')
    # print("Finish")

if __name__ == '__main__':
    main()
