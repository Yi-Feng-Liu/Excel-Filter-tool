import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, numbers
import copy
import json
import sys, os



class Judge_Metabolic_Syndrome:
    def __init__(self, io, src_worksheet, select_years, save_sheet_name, years_text=None, from_summary=False):
        self.io = io
        self.src_worksheet = src_worksheet
        self.dst_worksheet = save_sheet_name
        self.select_years = select_years
        self.years_text = years_text
        self.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.font = Font(name='Calibri', color='FF0000')
        self.font_type = Font(name='Calibri')
        self.from_summary = from_summary
        self.gender_dict = {'gender': 5}


    def resource_path(self, relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)


    def get_json_data(self):
        with open(self.resource_path("util/types.json"), encoding="utf-8") as f:
            data = json.load(f)
        return data


    def get_Metabolic_Syndrome_Dict(self):  
        data = self.get_json_data()
        Metabolic_Syndrome = data['Metabolic_Syndrome']
        column_dict = Metabolic_Syndrome['column_dict']
        standard_dict = Metabolic_Syndrome['Male_standard']
        return column_dict, standard_dict


    def change_date_time(self, worksheet, number_of_column):
        """Remove hours:minute:second format of the datetime 

        Args:
            worksheet : excel worksheet
            number_of_column : If column name is G, the number of column is 7, etc.

        Returns:
            _worksheet: 
        """
        # 變更時間格式
        for row in worksheet.iter_rows(min_row=2, min_col=number_of_column, max_col=number_of_column):
            for cell in row:
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
        return worksheet


    def place_center(self, worksheet):
        align = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = align
        return worksheet


    def set_specific_column_format(self, worksheet:str, eng_column:str, width=15, only_change_width=False):
        """set the specific column format.

        Args:
            worksheet (str): Excel worksheet
            eng_column (str): like 'A' or 'G' column
        """
        if only_change_width==True:
            worksheet.column_dimensions[eng_column].width = width
        else:
            worksheet.column_dimensions[eng_column].width = width
            worksheet[f'{eng_column}1'].fill = self.fill
            worksheet[f'{eng_column}1'].font = self.font
        return worksheet


    def change_font_color_format(self, cell):
        """Change font color 

        Args:
            cell : the cell coordinate
        """
        cell.font = self.font
    

    def change_font_type_format(self, worksheet):
        """Change font type 

        Args:
            cell : the cell coordinate
        """
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                # 設置儲存格的字體
                cell.font = self.font_type
        return worksheet


    def label_over_standard(self, worksheet):
        """Use to label the cell, if cell's value exceed the standard

        Args:
            worksheet: the excel work sheet 

        Returns:
            worksheet
        """
        worksheet = self.change_font_type_format(worksheet)
        column_dict, standard_dict = self.get_Metabolic_Syndrome_Dict()
        for row in worksheet.iter_rows(min_row=2):
            people_name = row[1]
            gender = row[self.gender_dict['gender']] 
            over_standard_cnt = 0
            if gender.value == '男':
                for key, value in column_dict.items():
                    if row[value].value is None:
                        continue
                    if row[value].value == '無資料':
                        continue
                    if isinstance(row[value].value, str):
                        row[value].value = float(row[value].value)
                        if key == 'hdlc' and row[value].value < standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)
                    else:
                        if key == 'waistline' and row[value].value >= standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key =='hdlc' and row[value].value < standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)

            elif gender.value == '女':
                for key, value in column_dict.items():
                    if row[value].value is None:
                        continue
                    if row[value].value == '無資料':
                        continue
                    if isinstance(row[value].value, str):
                        row[value].value = float(row[value].value.split('(')[0])
                        if key == 'waistline' and row[value].value >= standard_dict[key]-10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key =='hdlc' and row[value].value < standard_dict[key]+10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)
                    else:
                        if key == 'waistline' and row[value].value >= standard_dict[key]-10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key =='hdlc' and row[value].value < standard_dict[key]+10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)
        return worksheet
    

    def copy_title_format(self, ws1, ws2):
        for row in ws1.iter_rows(min_row=1, max_row=1):
            for cell in row:
                # copy cell
                ws2[cell.coordinate].font = copy.copy(cell.font)
                ws2[cell.coordinate].fill = copy.copy(cell.fill)
                ws2[cell.coordinate].border = copy.copy(cell.border)
                ws2[cell.coordinate].alignment = copy.copy(cell.alignment)
                ws2[cell.coordinate].number_format = copy.copy(cell.number_format)
                ws2.row_dimensions[cell.row].height = copy.copy(ws1.row_dimensions[cell.row].height)
                ws2.column_dimensions[cell.column_letter].width = copy.copy(ws1.column_dimensions[cell.column_letter].width)
        return ws2
    

    def copy_format_from_sheet1(self):
        """Copy the original sheet header format to specific sheet

        Including cell's fill, font, color, alignment, dimensions.
        """
        workbook = openpyxl.load_workbook(self.io)

        ws1 = workbook[self.src_worksheet]
        ws2 = workbook[self.dst_worksheet]
        # copy format sheet1 header to sheet2 header
        ws2 = self.copy_title_format(ws1=ws1, ws2=ws2)
        ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='U', width=13, only_change_width=False)
        if self.from_summary==True:
            start_column = 71 # G
            end_column = 81 #Q
            for i in range(start_column, end_column+1):
                ws2 = self.set_specific_column_format(worksheet=ws2, eng_column=chr(i), width=15, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='E', width=42, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='R', width=45, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='S', width=35, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='T', width=24, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='V', width=13)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='W', width=13)
        ws2 = self.change_date_time(worksheet=ws2, number_of_column=7)
        ws2 = self.place_center(worksheet=ws2)

        # label_over_standard worksheet
        # ws1 = self.label_over_standard(worksheet=ws1)
        # only process new sheet
        ws2 = self.label_over_standard(worksheet=ws2)
        workbook.save(self.io)
        print("saved")


    def read_file(self):
        if self.from_summary==False:
            df = pd.read_excel(self.io, sheet_name='健檢資料', engine='openpyxl')
        else:
            df = pd.read_excel(self.save_file_path, sheet_name='健檢資料', engine='openpyxl')
        return df
    
    
    def process_Metabolic_Syndrome(self, df:pd.DataFrame):
        """Filter file and add new column named 超過標準數,

        Args:
            df (pd.DataFrame): Original dataframe
        Returns:
            dataframe
        """
        df['超過標準數'] = 0
        column_dict, standard_dict = self.get_Metabolic_Syndrome_Dict()
        if self.from_summary==False:
            df = df[df['年度代碼'].str.startswith(self.select_years)]
        else:
            df = df[df['年度代碼'].str.startswith(self.years_text)]
             
        for i in range(len(df.index)):
            # name = df.iloc[i, 1] 
            gender = df.iloc[i, self.gender_dict['gender']] 
            over_standard_cnt = 0
            if gender == '男':
                for key, value in column_dict.items():
                    df_value = df.iloc[i, value]
                    if pd.isna(df_value):
                        continue
                    if df_value=='無資料':
                        df_value = ''
                        continue
                    if isinstance(df_value, str):
                        df_value = float(df_value)
                        if key == 'hdlc' and df_value < standard_dict[key]:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= standard_dict[key]:
                            over_standard_cnt += 1
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
                    else:
                        if key == 'hdlc' and df_value < standard_dict[key]:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= standard_dict[key]:
                            over_standard_cnt += 1
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
            elif gender == '女':
                for key, value in column_dict.items():
                    df_value = df.iloc[i, value]
                    if pd.isna(df_value):
                        continue
                    if df_value=='無資料':
                        df_value = ''
                        continue
                    if isinstance(df_value, str):
                        df_value = float(df_value.split('(')[0])
                        if key == 'waistline' and df_value >= standard_dict[key]-10:
                            over_standard_cnt += 1
                        elif key =='hdlc' and df_value < standard_dict[key]+10:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= standard_dict[key]:
                            over_standard_cnt += 1 
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
                    else:
                        if key == 'waistline' and df_value >= standard_dict[key]-10:
                            over_standard_cnt += 1
                        elif key =='hdlc' and df_value < standard_dict[key]+10:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= standard_dict[key]:
                            over_standard_cnt += 1 
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
        df = df.sort_values(by=['超過標準數'], ascending=False)
        return df
        
    
    def save_file_to_excel(self, df):
        # 建立一個新的 ExcelWriter 物件
        writer = pd.ExcelWriter(self.io, mode='a', engine='openpyxl', if_sheet_exists='replace')
        df.to_excel(writer, sheet_name=self.dst_worksheet, index=False)
        writer.close() 


    def main_procesdure(self):
        df = self.read_file()
        df = self.process_Metabolic_Syndrome(df)
        self.save_file_to_excel(df)
        self.copy_format_from_sheet1()


class Metabolic_Syndrome_From_Summary(Judge_Metabolic_Syndrome):
    def __init__(self, io, src_worksheet, save_sheet_name, years_text, from_summary=True):
        super().__init__(io, src_worksheet, save_sheet_name, years_text)
        self.io = io
        self.dst_worksheet = save_sheet_name
        self.years_text = years_text
        self.from_summary = from_summary


    def change_column_name(self, df, specific_column, goal_column_name):
        speific_df = df.copy()
        speific_df = speific_df[specific_column]
        speific_df_columns = speific_df.columns.to_list()
        for i in range(len(speific_df_columns)):
            speific_df.rename(columns={speific_df_columns[i]: goal_column_name[i]}, inplace=True)  
        return speific_df


    def set_column_name(self):
        self.df = pd.read_excel(self.io, engine='openpyxl')
        self.df['年度代碼'] = self.years_text
        self.df['部門代號'] = 'X001'
        self.df['健檢過程備註說明'] = ''
        self.df = self.df.drop(labels=0, axis=0)

        specific_column = ['年度代碼', '姓名', '工／學號', '部門代號', '部門／科系', '性別', '生日', '身高', '體重', '腰圍', '收縮壓', '舒張壓', 'AC飯前血糖', 'T-CHO總膽固醇', 'TG三酸甘油脂', 'HDL高密度脂蛋白', 'LDL低密度脂蛋白', '吸菸習慣', '既往病史', '健檢過程備註說明', 'SGOT血清麩酸草酸轉氨脢', 'SGPT血清麩酸丙銅轉氨脢']

        self.goal_column_name = ['年度代碼', '姓名', '員工編號', '部門代號', '部門名稱', '性別', '出生年月日', '身高_cm', '體重_kg', '腰圍_cm', '收縮壓', '舒張壓', '飯前血醣', '總膽固醇', '三酸甘油脂', '高密度膽固醇', '低密度膽固醇', '抽菸習慣', '既往病歷', '健檢過程備註說明', 'SGOT', 'SGOT']

        speific_df = self.change_column_name(self.df, specific_column, self.goal_column_name)
        return speific_df
    
    
    def main_procesdure(self):
        speific_df = self.set_column_name()
        speific_df = self.process_Metabolic_Syndrome(speific_df)
        self.save_file_to_excel(speific_df)
        self.copy_format_from_sheet1()

        
        

class Judge_Work_Pressure(Metabolic_Syndrome_From_Summary):
    def __init__(self, io, src_worksheet, save_sheet_name, years_text):
        super().__init__(io, src_worksheet, save_sheet_name, years_text)
        self.io = io
        self.src_worksheet = src_worksheet
        self.dst_worksheet = save_sheet_name
        self.years_text = years_text


    def __call__(self):
        self.insert_work_type_level_and_save()
        self.change_column_format_and_save()
        print('OK')


    def change_column_name(self, df:pd.DataFrame, goal_column_name:str):
        """The original column order should be same as goal column, this function only change the name of original column to goal.
        """
        speific_df = df.copy()
        speific_df_columns = speific_df.columns.to_list()
        for i in range(len(speific_df_columns)):
            speific_df.rename(columns={speific_df_columns[i]: goal_column_name[i]}, inplace=True)  
        return speific_df


    def get_work_pressure_dict(self):
        data = self.get_json_data()
        work_pressure_dict = data['Work_Pressure']
        self.tired_level = work_pressure_dict['tired_level']
        self.work_time_level = work_pressure_dict['work_time_level']
        self.sentimental_level = work_pressure_dict['sentimental_level']


    def read_file_and_change_col_name(self):
        self.df = pd.read_excel(self.io, sheet_name=self.src_worksheet, engine='openpyxl')
        self.df.drop(labels=['歸屬廠區', '單位', '工作班別'], axis=1, inplace=True)
        # add column & let's order same as template
        self.df.insert(0, '年度代碼', value=self.years_text)
        self.df.insert(9, 'I_Score', value=0)
        self.df.insert(10, 'I_Risk', value=0)
        self.df.insert(18, 'J_Score', value=0)
        self.df.insert(19, 'J_Risk', value=0)
        self.df.insert(22, '工作型態內容說明', value='無')
        self.df.insert(23, '工作負荷等級', value=0)

        target_col = ['年度代碼', '姓名', '員工編號', 'I01', 'I02', 'I03', 'I04', 'I05', 'I06', 'I_Score', 'I_Risk', 'J01', 'J02', 'J03', 'J04','J05', 'J06', 'J07', 'J_Score', 'J_Risk', '月加班時數等級', '工作型態評估等級', '工作型態內容說明', '工作負荷等級']
        self.df = self.change_column_name(self.df, target_col)
        return self.df


    def remove_empty_space(self, x):
        if isinstance(x, str):
            x = x.strip()
            return x
        elif isinstance(x, list):
            while "" in x:
                x.remove("")
            return x


    def procrss_name_col(self):
        self.df = self.read_file_and_change_col_name()
        self.df['姓名'] = self.df['姓名'].apply(lambda x: x.split(",")[-1])
        self.df['姓名'] = self.df['姓名'].apply(lambda x: self.remove_empty_space(x))
        return self.df


    def chage_series_text_to_value(self):
        self.get_work_pressure_dict()
        self.df = self.procrss_name_col()
        # change I columns value
        for i in self.df.iloc[:, 3:9]:
            series_i = self.df[i]
            self.df[i] = series_i.apply(lambda x: self.tired_level.get(x))
        # change J columns value
        for j in self.df.iloc[:, 11:14]:
            series_j = self.df[j]
            self.df[j] = series_j.apply(lambda x: self.sentimental_level.get(x))
        for j in self.df.iloc[:, 14:18]:
            series_j = self.df[j]
            self.df[j] = series_j.apply(lambda x: self.tired_level.get(x))
        # change 月加班時數等級 value
        self.df.iloc[:, 20] = self.df.iloc[:, 20].apply(lambda x: self.work_time_level.get(x))
 

    def calculate_level(self, df, *args):
        sum_series = 0
        for i in args:
            sum_series += (5 - df[i])
        series_value = (25*sum_series)/len(args)
        return series_value


    def insert_score_value(self):
        self.chage_series_text_to_value()
        I_Score_value = self.calculate_level(self.df, 'I01', 'I02', 'I03', 'I04', 'I05', 'I06')
        J_Score_value = self.calculate_level(self.df, 'J01', 'J02', 'J03', 'J04', 'J05', 'J06', 'J07')
        self.df['I_Score'] = I_Score_value.apply(lambda x: f'{x:.1f}').astype(float)
        self.df['J_Score'] = J_Score_value.apply(lambda x: f'{x:.1f}').astype(float)
        

    def insert_risk_value(self):
        self.insert_score_value()
        # insert value to I_Risk by score condition
        for i in range(len(self.df['I_Score'])):
            if self.df['I_Score'][i] <= 50:
                self.df.loc[i, 'I_Risk'] = 0
            elif self.df['I_Score'][i] >= 70:
                self.df.loc[i, 'I_Risk'] = 2
            else:
                self.df.loc[i, 'I_Risk'] = 1

        for j in range(len(self.df['J_Score'])):
            if self.df['J_Score'][j] <= 45:
                self.df.loc[j, 'J_Risk'] = 0
            elif self.df['J_Score'][j] >= 60:
                self.df.loc[j, 'J_Risk'] = 2
            else:
                self.df.loc[j, 'J_Risk'] = 1


    def insert_work_loading_level(self):
        self.insert_risk_value()
        # 插入工作型態評估等級, Condiction from , I & J Risk, 月加班時數等級
        for i in range(len(self.df.index)):
            I_Risk_value = self.df.loc[i, 'I_Risk']
            J_Risk_value = self.df.loc[i, 'J_Risk'] 
            work_time_level = self.df.loc[i, '月加班時數等級'] 
            loading_level = max(I_Risk_value, J_Risk_value, work_time_level)
            self.df.loc[i, '工作負荷等級'] = loading_level

    
    def process_work_type_to_level(self, x):
        x = x.split(";")
        for i in range(len(x)):
            if x[i] == '無以下特殊形態之工作':
                x[i] = ""
        x = self.remove_empty_space(x)
        return len(x)


    def judge_work_type_level(self, x):
        if x >= 0 and x <= 1:
            return 0
        elif x >= 2 and x <= 3:
            return 1
        elif x >= 4:
            return 2
        

    def insert_work_type_level_and_save(self):
        self.insert_work_loading_level()
        self.df['工作型態評估等級'] = self.df['工作型態評估等級'].apply(lambda x: self.process_work_type_to_level(x))
        self.df['工作型態評估等級'] = self.df['工作型態評估等級'].apply(lambda x: self.judge_work_type_level(x))
        # print(max(self.df['工作型態評估等級']))
        self.save_file_to_excel(self.df)
        print('Save Successfully')


    def set_specific_column_format(self, worksheet:str, eng_column:str, width=10, only_change_font_color=False, **kwargs):
        """set the specific column format.

        **kwargs:

        cell_color: Hexadecimal color, default FFC7CE

        font_name: str --> choose font type, default None

        font_color: Hexadecimal color, default FF0000

        only_change_font_color: bool
        """
        only_change_width = kwargs.get('only_change_width', False)
        cell_color = kwargs.get('cell_color', '00B0F0')
        font_color = kwargs.get('font_color', 'FFFFFF')
        font_name = kwargs.get('font_name', '微軟正黑體')

        if only_change_font_color == True:
            worksheet[f'{eng_column}1'].font = Font(name=font_name, color=font_color, bold=True)
        else:
            if only_change_width == True:
                worksheet.column_dimensions[eng_column].width = width
            else:
                worksheet.column_dimensions[eng_column].width = width
                worksheet[f'{eng_column}1'].fill = PatternFill(
                    start_color=cell_color, end_color=cell_color, fill_type='solid'
                )
                worksheet[f'{eng_column}1'].font = Font(name=font_name, color=font_color, bold=True)
        return worksheet


    def change_column_format_and_save(self):
        """Copy the original sheet header format to specific sheet

        Including cell's fill, font, color, alignment, dimensions.
        """
        workbook = openpyxl.load_workbook(self.io)

        ws2 = workbook[self.dst_worksheet]

        start_column = 68 # D
        end_column = 88 # X
        deep_color_cell = [74, 75, 83, 84]
        yellow_text_cell = [76, 77, 78, 79, 80, 81, 82, 83, 84]
        change_width_cell = [85, 86, 87]
        ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='A', width=10, cell_color='FFFF00', font_color='FF0000')
        ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='B', width=12, cell_color='008000')
        ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='C', width=10, cell_color='008000')
        for i in range(start_column, end_column+1):
            eng_col_name = chr(i)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column=eng_col_name)
            if i in deep_color_cell:
                ws2 = self.set_specific_column_format(worksheet=ws2, eng_column=eng_col_name, cell_color='0070C0')
            if i in yellow_text_cell:
                ws2 = self.set_specific_column_format(worksheet=ws2, eng_column=eng_col_name, only_change_font_color=True, font_color='FFFF00')
            if i in change_width_cell:
                ws2 = self.set_specific_column_format(worksheet=ws2, eng_column=eng_col_name, width=18, only_change_font_color=False, only_change_width=True)
            if i == 88:
                ws2 = self.set_specific_column_format(worksheet=ws2, eng_column=eng_col_name, width=20, cell_color='0070C0')
        
        ws2 = self.place_center(worksheet=ws2)
        workbook.save(self.io)
        print("saved")        



    

a = Judge_Work_Pressure('C:\\Users\\acer\\Desktop\\source.xlsx', '工作表1', 'test2', '112體檢')
a()